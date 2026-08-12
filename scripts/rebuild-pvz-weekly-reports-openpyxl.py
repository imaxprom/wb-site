#!/usr/bin/env python3
"""Rebuild the marked WB weekly reports as clean, standards-compliant XLSX files."""

from __future__ import annotations

import io
import json
import os
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
CASE_DIR = ROOT / "reports" / "pvz-case-43346319"
SOURCE_DIR = CASE_DIR / "source"
OUTPUT_DIR = CASE_DIR / "modern"
MANIFEST_PATH = CASE_DIR / "manifest.json"
OUTPUT_ARCHIVE = CASE_DIR / "wb-weekly-full-marked-box-43346319-modern-xlsx.zip"

EVIDENCE_COLUMNS = {
    11,  # K  Обоснование для оплаты
    28,  # AB Возмещение за выдачу и возврат товаров на ПВЗ
    35,  # AI Количество доставок
    36,  # AJ Количество возврата
    37,  # AK Услуги по доставке товара покупателю
    43,  # AQ Виды логистики, штрафов и корректировок ВВ
    44,  # AR Стикер МП
    50,  # AX Склад
    54,  # BB Номер сборочного задания
    57,  # BE Srid
    58,  # BF Возмещение издержек по перевозке/складским операциям
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TARGET_FILL = PatternFill("solid", fgColor="FFE699")
EVIDENCE_FILL = PatternFill("solid", fgColor="F4B183")
EVIDENCE_FONT = Font(bold=True, color="7F2704")
THIN_ALIGNMENT = Alignment(vertical="top")

COLUMN_WIDTHS = {
    "A": 9,
    "B": 15,
    "C": 20,
    "D": 17,
    "F": 24,
    "G": 42,
    "I": 18,
    "K": 32,
    "L": 15,
    "M": 15,
    "AB": 19,
    "AI": 14,
    "AJ": 14,
    "AK": 19,
    "AQ": 33,
    "AR": 18,
    "AX": 26,
    "BB": 21,
    "BE": 48,
    "BF": 24,
}


def extract_xlsx_buffer(report_id: int) -> bytes:
    archive_path = SOURCE_DIR / f"wb-weekly-{report_id}.zip"
    with zipfile.ZipFile(archive_path) as archive:
        xlsx_names = [name for name in archive.namelist() if name.lower().endswith(".xlsx")]
        if len(xlsx_names) != 1:
            raise RuntimeError(
                f"Report {report_id}: expected one XLSX, found {len(xlsx_names)}"
            )
        return archive.read(xlsx_names[0])


def styled_cell(ws, value, fill, font=None):
    cell = WriteOnlyCell(ws, value=value)
    cell.fill = fill
    if font is not None:
        cell.font = font
    cell.alignment = THIN_ALIGNMENT
    return cell


def create_legend_sheet(workbook: Workbook, report: dict, column_count: int) -> None:
    ws = workbook.create_sheet("Легенда")
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 95

    rows = [
        ("Поле", "Описание"),
        ("Отчёт WB", str(report["reportId"])),
        ("Период", report["period"]),
        ("Короб", "WB-MP-43346319"),
        ("Поставка", "WB-GI-246883602"),
        (
            "Жёлтая заливка",
            "Полная строка, относящаяся к одному из девяти спорных заказов.",
        ),
        (
            "Оранжевая заливка",
            "Ключевые поля: операция, логистика, направление, стикер, "
            "assembly ID, Srid и перевозочно-складские издержки.",
        ),
        (
            "Полнота",
            f"Все исходные строки и все {column_count} колонок отчёта WB сохранены.",
        ),
        (
            "Помеченные значения «№»",
            ", ".join(str(value) for value in report["sourceRows"]),
        ),
        (
            "Excel-строки",
            ", ".join(str(value) for value in report["spreadsheetRows"]),
        ),
    ]

    for row_index, values in enumerate(rows, start=1):
        cells = []
        for value in values:
            if row_index == 1:
                cells.append(styled_cell(ws, value, HEADER_FILL, HEADER_FONT))
            else:
                cell = WriteOnlyCell(ws, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cells.append(cell)
        ws.append(cells)


def rebuild_report(report: dict) -> Path:
    source_buffer = extract_xlsx_buffer(report["reportId"])
    source_workbook = load_workbook(
        io.BytesIO(source_buffer),
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    source_sheet = source_workbook.worksheets[0]
    # WB writes an incorrect A1 dimension despite a full worksheet.
    source_sheet.reset_dimensions()
    source_rows = source_sheet.iter_rows(values_only=True)
    headers = next(source_rows)
    source_column_count = len(headers)

    output_workbook = Workbook(write_only=True)
    create_legend_sheet(output_workbook, report, source_column_count)
    output_sheet = output_workbook.create_sheet("Полный отчёт WB")
    output_sheet.freeze_panes = "A2"
    output_sheet.sheet_view.showGridLines = True

    for column, width in COLUMN_WIDTHS.items():
        output_sheet.column_dimensions[column].width = width

    target_rows = set(report["spreadsheetRows"])
    row_count = 1
    column_count = source_column_count

    rendered_headers = []
    for column_index, value in enumerate(headers, start=1):
        fill = EVIDENCE_FILL if column_index in EVIDENCE_COLUMNS else HEADER_FILL
        font = EVIDENCE_FONT if column_index in EVIDENCE_COLUMNS else HEADER_FONT
        rendered_headers.append(styled_cell(output_sheet, value, fill, font))
    output_sheet.append(rendered_headers)

    for excel_row, values in enumerate(source_rows, start=2):
        row_count = excel_row
        column_count = max(column_count, len(values))

        if excel_row in target_rows:
            rendered = []
            for column_index, value in enumerate(values, start=1):
                if column_index in EVIDENCE_COLUMNS:
                    rendered.append(
                        styled_cell(output_sheet, value, EVIDENCE_FILL, EVIDENCE_FONT)
                    )
                else:
                    rendered.append(styled_cell(output_sheet, value, TARGET_FILL))
            output_sheet.append(rendered)
        else:
            output_sheet.append(values)

        if excel_row % 10000 == 0:
            print(
                f"  #{report['reportId']}: {excel_row - 1} rows streamed",
                flush=True,
            )

    source_workbook.close()

    expected_rows = report["dataRows"] + 1
    if row_count != expected_rows:
        raise RuntimeError(
            f"Report {report['reportId']}: expected {expected_rows} rows, got {row_count}"
        )
    if column_count != source_column_count:
        raise RuntimeError(
            f"Report {report['reportId']}: header has {source_column_count} columns, "
            f"data has {column_count}"
        )

    output_sheet.auto_filter.ref = (
        f"A1:{get_column_letter(column_count)}{row_count}"
    )
    output_sheet.sheet_properties.pageSetUpPr.fitToPage = True

    period_from, period_to = report["period"].split("—")
    output_path = (
        OUTPUT_DIR
        / f"WB_{report['reportId']}_{period_from}_{period_to}_modern.xlsx"
    )
    output_workbook.save(output_path)

    print(
        f"OK #{report['reportId']}: {row_count - 1} data rows, "
        f"{column_count} columns, {output_path.stat().st_size} bytes",
        flush=True,
    )
    return output_path


def build_readme(manifest: dict, outputs: list[Path]) -> str:
    lines = [
        "ПОЛНЫЕ ЕЖЕНЕДЕЛЬНЫЕ ОТЧЁТЫ WB — КОРОБ WB-MP-43346319",
        "",
        "Файлы заново сформированы в современном формате Office Open XML (.xlsx).",
        "Каждый файл содержит лист «Легенда» и лист «Полный отчёт WB».",
        "Все строки и все исходные колонки сохранены.",
        "",
        "Жёлтая заливка — строки девяти спорных заказов.",
        "Оранжевая заливка — ключевые поля логистики и идентификаторы.",
        "",
    ]
    for report, output in zip(manifest["fullReports"], outputs):
        lines.append(
            f"№{report['reportId']} · {report['period']} · "
            f"{report['dataRows']} строк · {output.name}"
        )
    return "\n".join(lines) + "\n"


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    manifest = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    outputs = [rebuild_report(report) for report in manifest["fullReports"]]
    readme = build_readme(manifest, outputs)
    readme_path = OUTPUT_DIR / "README_легенда.txt"
    readme_path.write_text(readme, encoding="utf-8")

    with zipfile.ZipFile(
        OUTPUT_ARCHIVE,
        "w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=6,
    ) as archive:
        archive.writestr("README_легенда.txt", readme)
        for output in outputs:
            archive.write(output, arcname=output.name)

    print(
        f"ARCHIVE {OUTPUT_ARCHIVE} {OUTPUT_ARCHIVE.stat().st_size} bytes",
        flush=True,
    )


if __name__ == "__main__":
    main()
